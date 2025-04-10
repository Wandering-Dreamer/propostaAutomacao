import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from openpyxl.utils import get_column_letter
import csv
from openpyxl import Workbook
import docx
from docx.shared import Pt, Mm, Cm, Inches
from docx.enum.table import WD_ALIGN_VERTICAL, WD_TABLE_ALIGNMENT
from docx.oxml.ns import nsdecls
from docx.oxml import parse_xml

df = pd.read_excel("SA38 Test.xlsx")
unique_values = df.iloc[:, 16].drop_duplicates().index

df2 = df.iloc[1:15, [2, 13]]
df2.columns = ['Serial Number', 'Net Value']
#print(df2.groupby(by=["Serial Number"]).sum())
df3 = df2.groupby(by=["Serial Number"], sort=False).sum()

new = df2.groupby(["Serial Number"], as_index=False, sort=False).sum()
#df2 = df2.drop_duplicates(subset=['Serial Number'])
df2 = df2.assign(Count = df2.groupby('Serial Number')['Net Value'].transform('sum'))
print(df2)


