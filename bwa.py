import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Sample data in pandas DataFrame
data = {
    'Name': ['John', 'Alice', 'Bob'],
    'Age': [30, 25, 28],
    'Location': ['New York', 'London', 'Paris']
}
df = pd.DataFrame(data)

# Write the DataFrame to Excel
file_path = 'output.xlsx'
df.to_excel(file_path, index=False)

# Load the Excel workbook using openpyxl
wb = load_workbook(file_path)
ws = wb.active

# Determine the range of cells to merge and center (e.g., A1 to C1)
start_row = 1
end_row = 1
start_column = 1
end_column = len(df.columns)

# Merge and center the cells
merge_range = f"{ws.cell(row=start_row, column=start_column).coordinate}:{ws.cell(row=end_row, column=end_column).coordinate}"
ws.merge_cells(merge_range)

# Save the changes
wb.save(file_path)